#!/usr/bin/env python3
"""
STEP 5 — Inter-rater agreement + cluster distribution (run AFTER both coders finish).

Reads the two coded workbooks, merges on act_id, and reports:
  - Cohen's kappa and percent agreement
  - a 9x9 confusion matrix (IRE RE BI CO RD EI PC GD ND)
  - the list of disagreements
  - the cluster distribution (counts and % per SEDA cluster) = draft Table 3

Prints a summary and writes seda/results/agreement_summary.md.

Usage:
    python seda/analyze_agreement.py \
        [--a seda/coding/coder_A.xlsx] [--b seda/coding/coder_B.xlsx]
"""
import os
import sys
import argparse
from collections import Counter, defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

CODES = ["IRE", "RE", "BI", "CO", "RD", "EI", "PC", "GD", "ND"]
OUT = "seda/results/agreement_summary.md"


def _read_coding(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Coding"]
    header = [c.value for c in ws[1]]
    ci = {name: i for i, name in enumerate(header)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        act_id = row[ci["act_id"]]
        if not act_id:
            continue
        code = (row[ci["code"]] or "").strip().upper() if row[ci["code"]] else ""
        out[act_id] = {
            "code": code,
            "act_text": row[ci["act_text"]],
            "error_category": row[ci["error_category"]],
            "script_id": row[ci["script_id"]],
        }
    return out


def cohen_kappa(pairs):
    """pairs = list of (code_a, code_b). Returns (kappa, percent_agreement, n)."""
    n = len(pairs)
    if n == 0:
        return float("nan"), float("nan"), 0
    agree = sum(1 for a, b in pairs if a == b)
    po = agree / n
    ca = Counter(a for a, _ in pairs)
    cb = Counter(b for _, b in pairs)
    pe = sum((ca[k] / n) * (cb[k] / n) for k in set(ca) | set(cb))
    kappa = (po - pe) / (1 - pe) if (1 - pe) else float("nan")
    return kappa, po, n


def confusion(pairs, labels):
    m = defaultdict(lambda: defaultdict(int))
    for a, b in pairs:
        m[a][b] += 1
    return m


def _md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for r in rows:
        out.append("| " + " | ".join(str(x) for x in r) + " |")
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--a", default="seda/coding/coder_A.xlsx")
    ap.add_argument("--b", default="seda/coding/coder_B.xlsx")
    args = ap.parse_args()

    A, B = _read_coding(args.a), _read_coding(args.b)
    common = [k for k in A if k in B]
    coded = [k for k in common if A[k]["code"] and B[k]["code"]]
    uncoded = [k for k in common if not (A[k]["code"] and B[k]["code"])]

    pairs = [(A[k]["code"], B[k]["code"]) for k in coded]
    labels = CODES + sorted({c for p in pairs for c in p if c not in CODES})

    kappa, po, n = cohen_kappa(pairs)

    # confusion matrix
    m = confusion(pairs, labels)
    conf_rows = []
    used = [l for l in labels if any(m[l].values()) or any(m[x][l] for x in labels)]
    for a in used:
        conf_rows.append([a] + [m[a][b] for b in used])

    # disagreements
    disagree = [(k, A[k]["code"], B[k]["code"], A[k]["error_category"], A[k]["act_text"])
                for k in coded if A[k]["code"] != B[k]["code"]]

    # cluster distribution — union across both coders on agreed acts + each coder's marginal
    dist_a = Counter(A[k]["code"] for k in coded)
    dist_b = Counter(B[k]["code"] for k in coded)
    # "consensus" distribution: use agreed code; for disagreements count as unresolved
    consensus = Counter(A[k]["code"] for k in coded if A[k]["code"] == B[k]["code"])
    total_consensus = sum(consensus.values())

    lines = []
    lines.append("# SEDA Inter-Rater Agreement & Cluster Distribution\n")
    lines.append(f"- Acts common to both files: **{len(common)}**")
    lines.append(f"- Acts coded by BOTH: **{len(coded)}**")
    if uncoded:
        lines.append(f"- ⚠️ Acts still uncoded by at least one coder: **{len(uncoded)}** (excluded)")
    lines.append(f"- **Percent agreement:** {po*100:.1f}%")
    lines.append(f"- **Cohen's kappa:** {kappa:.3f}")
    lines.append("")

    lines.append("## Confusion matrix (rows = Coder A, cols = Coder B)\n")
    lines.append(_md_table(["A\\B"] + used, conf_rows))
    lines.append("")

    lines.append(f"## Cluster distribution (consensus on {total_consensus} agreed acts) — draft Table 3\n")
    dist_rows = []
    for code in [c for c in CODES if consensus.get(c)] + \
               sorted(k for k in consensus if k not in CODES):
        cnt = consensus[code]
        pct = 100 * cnt / total_consensus if total_consensus else 0
        dist_rows.append([code, cnt, f"{pct:.1f}%"])
    lines.append(_md_table(["SEDA cluster", "count", "%"], dist_rows))
    lines.append("")
    lines.append("### Per-coder marginals (all coded acts)\n")
    marg_rows = [[c, dist_a.get(c, 0), dist_b.get(c, 0)]
                 for c in CODES if dist_a.get(c) or dist_b.get(c)]
    lines.append(_md_table(["cluster", "Coder A", "Coder B"], marg_rows))
    lines.append("")

    lines.append(f"## Disagreements ({len(disagree)})\n")
    if disagree:
        drows = [[k, a, b, ec, (txt or "")[:80]] for k, a, b, ec, txt in disagree]
        lines.append(_md_table(["act_id", "A", "B", "error_category", "act_text (trunc)"], drows))
    else:
        lines.append("_None — perfect agreement._")
    lines.append("")

    report = "\n".join(lines)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write(report)

    # paste-ready block for PAPER_CORRECTIONS.md §4.7 placeholders
    snip = ["<!-- paste into PAPER_CORRECTIONS.md §4.7 'Results — FILL AFTER CODING' -->",
            f"- **Inter-rater reliability (RQ3):** Cohen's κ = {kappa:.2f} "
            f"(percent agreement {po*100:.1f}%), _n_ = {len(coded)} acts coded by both.",
            "",
            "| SEDA cluster | count | % |",
            "|---|---|---|"]
    for code in CODES:
        cnt = consensus.get(code, 0)
        pct = 100 * cnt / total_consensus if total_consensus else 0
        snip.append(f"| {code} | {cnt} | {pct:.1f}% |")
    snip_path = "seda/results/paper_4_7_snippet.md"
    with open(snip_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(snip) + "\n")

    # console summary
    print(f"n(coded by both) = {len(coded)}")
    print(f"percent agreement = {po*100:.1f}%")
    print(f"Cohen's kappa = {kappa:.3f}")
    print(f"disagreements = {len(disagree)}")
    print(f"Wrote {OUT}")
    print(f"Wrote {snip_path} (paste-ready §4.7 block)")


if __name__ == "__main__":
    main()
