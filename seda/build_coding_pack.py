#!/usr/bin/env python3
"""
STEP 4 — Build the two-coder pack (Excel) from the fixed units.

Produces:
  seda/coding/coder_A.xlsx   (blind)
  seda/coding/coder_B.xlsx   (blind)
  seda/coding/_merge_key.csv (PRIVATE — act_id -> canonical order; do NOT give to coders)

Each workbook has a "Codebook" sheet (8 SEDA clusters + ND, definitions, decision
rules, English examples, instructions) and a "Coding" sheet with a data-validated
`code` dropdown left EMPTY (blind). Row order is randomised independently per coder.

Usage:
    python seda/build_coding_pack.py
"""
import os
import csv
import sys
import random

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

UNITS = "seda/corpus/units.csv"
CODER_A = "seda/coding/coder_A.xlsx"
CODER_B = "seda/coding/coder_B.xlsx"
MERGE_KEY = "seda/coding/_merge_key.csv"

CODES = ["IRE", "RE", "BI", "CO", "RD", "EI", "PC", "GD", "ND"]

# SEDA clusters (Hennessy et al., 2016) + ND. (code, name, definition, decision rule, example)
CODEBOOK = [
    ("IRE", "Invite elaboration or reasoning",
     "Ask the student to explain, justify, build on, or expand their thinking.",
     "The act is a question/prompt asking WHY, HOW, or for more detail/justification.",
     "Ask the student to explain why they chose the past tense here."),
    ("RE", "Make reasoning explicit",
     "State, explain, or spell out reasoning, an account, or a justification.",
     "The act gives the logic/because behind an idea (not just the idea itself).",
     "Show how the topic sentence signals the paragraph's main idea."),
    ("BI", "Build on ideas",
     "Build on, elaborate, clarify, or refine one's own or the student's contribution.",
     "The act extends/reworks a prior idea rather than introducing a new one.",
     "Take the student's point and add a concluding sentence that ties it back to the thesis."),
    ("CO", "Connect",
     "Make links between ideas, texts, experiences, or prior learning.",
     "The act links the content to real life, another text, or earlier work.",
     "Relate the poem's theme of loss to a personal experience the student has had."),
    ("RD", "Reflect on dialogue or activity",
     "Evaluate or reflect on the process of learning, talking, or the activity itself.",
     "The act comments on HOW the learning/task is going (meta-level).",
     "Ask the student which step of the summary they found hardest and why."),
    ("EI", "Express or invite ideas",
     "Offer or ask for a relevant idea, opinion, or contribution.",
     "The act invites/states an idea or opinion WITHOUT requiring justification.",
     "Ask the student what they think the writer's main message is."),
    ("PC", "Positioning and coordination",
     "Express or invite agreement/disagreement; coordinate or challenge viewpoints.",
     "The act takes or asks for a stance relative to another view (agree/challenge/concede).",
     "Ask whether the student agrees with the author's argument, and why."),
    ("GD", "Guide direction of dialogue or activity",
     "Direct or focus the dialogue/activity; propose a way forward or next step.",
     "The act manages/steers the task — instructions, focusing, sequencing moves.",
     "Let's work through one simpler example together, then you try one on your own."),
    ("ND", "Non-dialogic / other",
     "Performs none of the dialogic functions above.",
     "Pure praise, greeting, logistics, or a bare instruction with no dialogic scaffolding.",
     "Well done, keep up the good effort."),
]

INSTRUCTIONS = [
    "Assign exactly ONE code per act (one row = one communicative act).",
    "Code the DOMINANT function of the act if more than one seems present.",
    "The units are FIXED — do not split, merge, or re-segment the acts.",
    "Use ND only when no dialogic cluster (IRE-GD) applies.",
    "Leave 'notes' for anything you want the other coder / adjudicator to see.",
    "Do not consult the other coder while coding (blind double-coding).",
]

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def _read_units():
    with open(UNITS, encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _codebook_sheet(ws):
    ws.title = "Codebook"
    ws["A1"] = "SEDA Coding Scheme — 8 clusters + ND"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    hdr = ["Code", "Cluster name", "Definition", "Decision rule", "English example"]
    ws.append(hdr)
    for c in ws[ws.max_row]:
        c.fill, c.font = HEAD_FILL, HEAD_FONT
    for code, name, defn, rule, ex in CODEBOOK:
        ws.append([code, name, defn, rule, ex])
    ws.append([])
    ws.append(["Instructions:"])
    ws[ws.max_row][0].font = Font(bold=True)
    for line in INSTRUCTIONS:
        ws.append([f"• {line}"])
    widths = [8, 32, 46, 46, 52]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=4, max_row=3 + len(CODEBOOK), min_col=3, max_col=5):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")


def _coding_sheet(wb, rows, seed):
    ws = wb.create_sheet("Coding")
    hdr = ["script_id", "topic", "error_category", "act_no", "act_id",
           "act_text", "code", "notes"]
    ws.append(hdr)
    for c in ws[1]:
        c.fill, c.font = HEAD_FILL, HEAD_FONT

    shuffled = list(rows)
    random.Random(seed).shuffle(shuffled)
    for r in shuffled:
        ws.append([r["script_id"], r["topic"], r["error_category"], int(r["act_no"]),
                   r["act_id"], r["act_text"], "", ""])

    # blind dropdown on `code` (col G)
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(CODES), allow_blank=True)
    dv.error = "Pick one of: " + ", ".join(CODES)
    dv.prompt = "Assign ONE SEDA code (see Codebook sheet)."
    ws.add_data_validation(dv)
    last = ws.max_row
    dv.add(f"G2:G{last}")

    widths = [10, 26, 22, 8, 12, 70, 10, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=last, min_col=6, max_col=6):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


def _build(path, rows, seed):
    wb = openpyxl.Workbook()
    _codebook_sheet(wb.active)
    _coding_sheet(wb, rows, seed)
    wb.save(path)


def main():
    rows = _read_units()

    # canonical order = the order in units.csv
    with open(MERGE_KEY, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["canonical_order", "act_id", "script_id", "error_category"])
        for i, r in enumerate(rows):
            w.writerow([i, r["act_id"], r["script_id"], r["error_category"]])

    _build(CODER_A, rows, seed=17)
    _build(CODER_B, rows, seed=83)

    print(f"Wrote {CODER_A} and {CODER_B} ({len(rows)} acts each, blind, independently shuffled)")
    print(f"Wrote private {MERGE_KEY} (do NOT share with coders)")


if __name__ == "__main__":
    main()
