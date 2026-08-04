#!/usr/bin/env python3
"""
STEP 4b — Automated SEDA coding by two INDEPENDENT LLM coders.

This does NOT fabricate a distribution. It sends each real communicative act
(seda/corpus/units.csv, 167 acts) to an LLM and records the model's own single
best-fitting SEDA cluster label. Two different model FAMILIES code the corpus
blind to each other, so analyze_agreement.py yields a genuine Cohen's kappa
between two automated coders — not a hand-picked number.

  Coder A  -> Cerebras  gpt-oss-120b
  Coder B  -> DeepSeek  deepseek-chat

Each coder is given the same codebook (definitions + decision rules) and must
return, for a batch of acts, strict JSON mapping act_id -> code. Codes outside
the 9-label set are recorded verbatim (analyze_agreement handles unknowns).

Usage:
    python3 seda/code_acts.py --coder A     # fills coder_A.xlsx  (Cerebras)
    python3 seda/code_acts.py --coder B     # fills coder_B.xlsx  (DeepSeek)
"""
import os
import csv
import json
import argparse

import time

import openpyxl
from openai import OpenAI, RateLimitError
from dotenv import load_dotenv

load_dotenv(override=True)

HERE = os.path.dirname(os.path.abspath(__file__))
UNITS = os.path.join(HERE, "corpus", "units.csv")
CODES = ["IRE", "RE", "BI", "CO", "RD", "EI", "PC", "GD", "ND"]
BATCH = 15

CODER = {
    "A": {
        "book": os.path.join(HERE, "coding", "coder_A.xlsx"),
        "client": lambda: OpenAI(base_url="https://api.groq.com/openai/v1",
                                 api_key=os.getenv("GROQ_API_KEY"), timeout=90.0),
        "model": "llama-3.3-70b-versatile",
    },
    "B": {
        "book": os.path.join(HERE, "coding", "coder_B.xlsx"),
        "client": lambda: OpenAI(base_url="https://api.deepseek.com/v1",
                                 api_key=os.getenv("DEEPSEEK_API_KEY"), timeout=90.0),
        "model": "deepseek-chat",
    },
}

CODEBOOK = """SEDA coding scheme — assign exactly ONE code per act (its DOMINANT function):

IRE  Invite elaboration or reasoning — asks the student WHY/HOW or to expand/justify.
RE   Make reasoning explicit — spells out the logic / the 'because' behind an idea.
BI   Build on ideas — extends, refines, or reworks an existing idea (own or student's).
CO   Connect — links content to real life, another text, or prior learning.
RD   Reflect on dialogue/activity — comments on HOW the learning/task is going (meta).
EI   Express or invite ideas — offers/asks for an idea or opinion, no justification required.
PC   Positioning & coordination — takes/asks for a stance on a view (agree/disagree/challenge).
GD   Guide direction — steers the task: instructions, next step, focusing, sequencing.
ND   Non-dialogic/other — bare praise, greeting, pure logistics; no dialogic scaffolding.

Rules: one code per act; code the dominant function if two seem present; the acts are
fixed (do not re-segment); use ND only when no IRE-GD cluster fits. Scripts mix English
and Bahasa Malaysia — code BM acts the same way. Some acts describe a teacher activity/
technique (e.g. a chart or worksheet); code by the dialogic function it performs."""


def _load_units():
    with open(UNITS, encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _classify(client, model, batch):
    listing = "\n".join(f'{u["act_id"]}: {u["act_text"]}' for u in batch)
    prompt = (
        CODEBOOK
        + "\n\nCode each communicative act below. Return ONLY a JSON object mapping "
        + 'each act_id to its single code, e.g. {"S001a01": "IRE"}. No prose.\n\n'
        + listing
    )
    for attempt in range(6):
        try:
            r = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You are an expert educational-dialogue "
                     "coder applying the SEDA scheme. Respond with strict JSON only."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.0,
                response_format={"type": "json_object"},
            )
            break
        except RateLimitError:
            wait = 5 * (attempt + 1)
            print(f"  429 rate-limited; backing off {wait}s")
            time.sleep(wait)
    else:
        raise SystemExit("persistent 429 after 6 retries")
    txt = r.choices[0].message.content.strip()
    data = json.loads(txt)
    if isinstance(data, list):
        data = data[0]
    return {k: str(v).strip().upper() for k, v in data.items()}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--coder", required=True, choices=["A", "B"])
    args = ap.parse_args()
    cfg = CODER[args.coder]
    client = cfg["client"]()

    units = _load_units()
    codes = {}
    for i in range(0, len(units), BATCH):
        batch = units[i:i + BATCH]
        got = _classify(client, cfg["model"], batch)
        codes.update(got)
        print(f"coder {args.coder} [{cfg['model']}]: coded {len(codes)}/{len(units)}")

    # retry any acts the model omitted from a batch, one at a time
    for _ in range(3):
        missing = [u for u in units if u["act_id"] not in codes]
        if not missing:
            break
        print(f"coder {args.coder}: retrying {len(missing)} omitted acts individually")
        for u in missing:
            codes.update(_classify(client, cfg["model"], [u]))
    missing = [u["act_id"] for u in units if u["act_id"] not in codes]
    if missing:
        raise SystemExit(f"coder {args.coder}: {len(missing)} acts uncoded: {missing[:5]}")

    # write into the Coding sheet, keyed by act_id (row order differs per file)
    wb = openpyxl.load_workbook(cfg["book"])
    ws = wb["Coding"]
    header = [c.value for c in ws[1]]
    ai, ci = header.index("act_id"), header.index("code")
    n = 0
    for row in ws.iter_rows(min_row=2):
        aid = row[ai].value
        if aid in codes:
            row[ci].value = codes[aid]
            n += 1
    wb.save(cfg["book"])
    print(f"coder {args.coder}: wrote {n} codes -> {cfg['book']}")


if __name__ == "__main__":
    main()
